import time
import logging
from fastapi import FastAPI, HTTPException
from datetime import datetime
from isoweek import Week
import openpyxl
from fastapi.middleware.cors import CORSMiddleware  # Импортируем CORS

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI()


app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Разрешить запросы с любых источников (лучше настроить конкретные источники)
    allow_credentials=True,
    allow_methods=["*"],  # Разрешить любые методы (GET, POST и т.д.)
    allow_headers=["*"],  # Разрешить любые заголовки
)

groups = [
    'ТОР-23', 'РЭГ-23', 'СЭЗС-23', 'ПР-23', 'ОПИ-23', 'ДПИ-23', 'МД-23/1', 'МД-23/2',
    'ИСИП-23/1', 'ИСИП-23/2', 'БУ-23', 'БД-23', 'Ф-23', 'ЗИМ-23', 'ЮР-23/1', 'ЮР-23/2',
    'ПКД-23', 'ТОР-22', 'РЭГ-22', 'КИП-22', 'СЭЗС-22', 'ПР-22', 'ОПИ-22', 'ДПИ-22', 'МД-22',
    'ПО-22/1', 'ПО-22/2', 'БУ-22', 'БД-22', 'Ф-22', 'ЗИМ-22', 'ПСО-22/1', 'ПСО-22/2', 'ПКД-22',
    'ТОР-21', 'РЭГ-21', 'КИП-21', 'СЭЗС-21', 'ПР-21', 'ОПИ-21', 'ДПИ-21', 'МД-21', 'ПО-21',
    'БУ-21', 'Ф-21', 'ЗИМ-21', 'ПСО-21/1', 'ПСО-21/2', 'ПКД-21', 'ТОР-20', 'РЭГ-20', 'СЭЗС-20',
    'ПР-20', 'ОПИ-20', 'ДПИ-20', 'БУ-20', 'МД-20', 'ПО-20', 'ПКД-20'
]


loaded_schedules = {
    "odd": None,
    "even": None
}


def get_lesson_number(i):
    return str(i+1)


def load_schedule_file(filename):
    """Загружает и возвращает активный лист из файла Excel."""
    try:
        wb_schedule = openpyxl.load_workbook(filename)
        return wb_schedule.active
    except FileNotFoundError:
        logger.error(f"Файл {filename} не найден.")
        raise HTTPException(status_code=404, detail="Schedule file not found")
    except Exception as e:
        logger.error(f"Ошибка при загрузке файла {filename}: {e}")
        raise HTTPException(status_code=500, detail="Internal Server Error")


def initialize_schedules():
    """Загружает расписания при запуске приложения."""
    loaded_schedules["odd"] = load_schedule_file("rasp_cet.xlsx")
    loaded_schedules["even"] = load_schedule_file("rasp_necet.xlsx")


@app.on_event("startup")
async def startup_event():
    """Функция инициализации, вызываемая при старте приложения."""
    initialize_schedules()
    print("Загрузка расписания выполнена успешно")


def get_schedule_for_day(user_group: str, day: str):
    """Получает расписание на указанный день для заданной группы."""
    start_time = time.perf_counter()
    today = datetime.today()
    current_day = today.weekday()
    current_week = Week.withdate(today).week

    if user_group not in groups:
        raise HTTPException(status_code=404, detail="Group not found")

    schedule_column = groups.index(user_group) + 3

    schedule_type = "even" if current_week % 2 == 0 else "odd"

    if current_day == 6:  # Если сегодня воскресенье
        schedule_type = "odd" if current_week % 2 != 0 else "even"

    sheet_schedule = loaded_schedules[schedule_type]

    if day == "tomorrow":
        target_day = (current_day + 1) % 7  # Завтрашний день
    elif day == "today":
        target_day = current_day  # Сегодня
    else:
        raise HTTPException(status_code=400, detail="Invalid day specified")

    start_row = 6 + (target_day * 6)
    end_row = start_row + 5

    schedule = {}
    for i in range(start_row, end_row + 1):
        lesson = sheet_schedule.cell(row=i, column=schedule_column).value
        cleaned_lesson = " ".join(lesson.split()) if lesson else "Нет"
        schedule[get_lesson_number(i - start_row)] = cleaned_lesson

    end_time = time.perf_counter()
    logger.info(f"Время выполнения запроса для группы {user_group} на {day}: {end_time - start_time:.10f} секунд")
    return schedule


def get_schedule_for_week(user_group: str, parity: str):
    """Получает расписание на неделю для заданной группы, структурированное по дням недели."""
    start_time = time.perf_counter()

    if user_group not in groups:
        raise HTTPException(status_code=404, detail="Group not found")

    if parity not in ["odd", "even"]:
        raise HTTPException(status_code=400, detail="Invalid parity specified")

    sheet_schedule = loaded_schedules[parity]

    schedule_column = groups.index(user_group) + 3

    week_schedule = {
        "pn": {},   # понедельник
        "vt": {},   # вторник
        "sr": {},   # среда
        "cht": {},  # четверг
        "pt": {},   # пятница
        "sb": {}    # суббота
    }

    for day_index, day_name in enumerate(week_schedule.keys()):
        start_row = 6 + (day_index * 6)
        end_row = start_row + 5

        for i in range(start_row, end_row + 1):
            lesson = sheet_schedule.cell(row=i, column=schedule_column).value
            cleaned_lesson = " ".join(lesson.split()) if lesson else "Нет"
            week_schedule[day_name][get_lesson_number(i - start_row)] = cleaned_lesson

    end_time = time.perf_counter()
    logger.info(f"Время выполнения запроса для группы {user_group} на {parity}: {end_time - start_time:.10f} секунд")

    return week_schedule


@app.get("/schedule/")
async def read_schedule(day: str, group: str):
    return get_schedule_for_day(group, day)


@app.get("/schedule/for_week/")
async def schedule_for_week(group: str, parity: str):
    return get_schedule_for_week(group, parity)
